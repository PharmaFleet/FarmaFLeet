import csv
import io
import logging
from typing import BinaryIO, Dict, List

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

CSV_ENCODINGS = ["utf-8", "utf-8-sig", "latin-1", "cp1256", "cp1252"]


class ExcelService:
    def parse_file(self, file: BinaryIO, filename: str = "") -> List[Dict]:
        """Parse an Excel/CSV/HTML file and return a list of dictionaries."""
        errors: Dict[str, str] = {}

        # CSV path - try multiple encodings
        if filename.lower().endswith(".csv"):
            for encoding in CSV_ENCODINGS:
                try:
                    text = file.read().decode(encoding)
                    reader = csv.DictReader(io.StringIO(text))
                    rows = list(reader)
                    if not rows:
                        raise ValueError("CSV file is empty")
                    logger.info(
                        f"CSV parsed with encoding={encoding}, rows={len(rows)}"
                    )
                    return rows
                except Exception as e:
                    errors[f"csv({encoding})"] = str(e)
                    file.seek(0)
            logger.error(f"CSV parse failed for '{filename}': {errors}")
            raise ValueError(
                f"CSV parse failed: {errors.get('csv(utf-8)', errors.get('csv(latin-1)', 'unknown'))}"
            )

        # Try .xlsx via openpyxl
        try:
            rows = self._read_xlsx(file)
            if rows:
                return rows
            errors["xlsx"] = "Empty workbook or no data rows"
            file.seek(0)
        except Exception as e:
            errors["xlsx"] = str(e)
            file.seek(0)

        # Try CSV (for files without .csv extension)
        for encoding in CSV_ENCODINGS:
            try:
                text = file.read().decode(encoding)
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
                if not rows:
                    raise ValueError("CSV file is empty")
                return rows
            except Exception:
                file.seek(0)

        # Try HTML table (Dynamics 365 exports)
        try:
            rows = self._read_html_table(file)
            if rows:
                return rows
            errors["html"] = "No tables found or table is empty"
        except Exception as e:
            errors["html"] = str(e)
            file.seek(0)

        logger.error(f"All parse attempts failed for '{filename}': {errors}")
        raise ValueError("Unsupported file format. Tried: xlsx, csv, html")

    def _read_xlsx(self, file: BinaryIO) -> List[Dict]:
        """Read an xlsx file using openpyxl and return list of dicts."""
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return []

        # Clean headers: convert to string, strip whitespace
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]

        result = []
        for row in rows_iter:
            record = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = val
            # Skip completely empty rows
            if any(v is not None for v in row):
                result.append(record)

        wb.close()
        return result

    def _read_html_table(self, file: BinaryIO) -> List[Dict]:
        """Read the first HTML table using lxml and return list of dicts."""
        from lxml import html as lxml_html

        content = file.read()
        doc = lxml_html.fromstring(content)
        tables = doc.xpath("//table")
        if not tables:
            return []

        table = tables[0]
        rows = table.xpath(".//tr")
        if not rows:
            return []

        # Extract headers from first row
        header_row = rows[0]
        headers = [
            (cell.text_content().strip() or f"col_{i}")
            for i, cell in enumerate(header_row.xpath(".//th|.//td"))
        ]

        result = []
        for row in rows[1:]:
            cells = row.xpath(".//td")
            record = {}
            for i, cell in enumerate(cells):
                if i < len(headers):
                    text = cell.text_content().strip()
                    record[headers[i]] = text if text else None
            if any(v is not None for v in record.values()):
                result.append(record)

        return result

    def read_local_file(self, file_path: str) -> List[Dict]:
        """Read a local Excel file."""
        with open(file_path, "rb") as f:
            return self._read_xlsx(f)

    @staticmethod
    def write_xlsx(data: List[Dict], sheet_name: str = "Sheet1") -> io.BytesIO:
        """Write a list of dicts to an xlsx BytesIO stream."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            stream = io.BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream

        headers = list(data[0].keys())
        ws.append(headers)

        for record in data:
            ws.append([record.get(h) for h in headers])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream


excel_service = ExcelService()
